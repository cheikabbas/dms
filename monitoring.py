import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

from messageBoxes import show_info_messagebox
import datetime


class Monitoring:
    def __init__(self, idvar, start, end, enq, vars, data, path):
        self.outputPath = path
        self.timeissues = {}
        self.id = {}
        self.enumeratorissues = {}
        self.idvar = idvar
        self.start = start
        self.end = end
        self.enq = enq
        self.vars = vars
        self.data = data
        self.convertdate()
        self.monitore()

    def duplicatedid(self):
        if self.idvar is not None:
            self.id['duplicates'] = self.data[self.data.duplicated(subset=[self.idvar], keep=False)]

    def missingid(self):
        if self.idvar is not None:
            nb_na = self.data.isnull().sum()
            pct_na = round(self.data.isna().mean() * 100, 2)
            self.id['missing'] = pd.DataFrame({"NB Valeurs manquantes": nb_na, "% Valeurs manquantes": pct_na})
            self.id['missing'].insert(loc=0, column='Variable', value=self.id['missing'].index)

    def end_before_start(self):
        mask = self.data[self.start] >= self.data[self.end]
        self.timeissues['end_before_start'] = self.data[mask]

    def lendayofsurvey(self):
        mask = self.data['Durée en mins'] / 60 > 24
        self.timeissues['onedaymore_duration'] = self.data[mask]

    def convertdate(self):
        self.data[self.start] = pd.to_datetime(self.data[self.start])
        self.data[self.end] = pd.to_datetime(self.data[self.end])
        self.data['Durée en mins'] = (self.data[self.end] - self.data[self.start]).dt.total_seconds() / 60

    def enumerators(self):
        self.enumeratorissues['duration'] = self.data.groupby(self.enq).agg(
            {'Durée en mins': ['count', 'mean', 'min', 'max']})

    def export(self):
        workbook = openpyxl.Workbook()

        # Missings
        workbook.create_sheet("Id manquants")
        worksheet = workbook["Id manquants"]
        if self.id['missing'] is not None:
            for row in dataframe_to_rows(self.id['missing']):
                worksheet.append(row)
        worksheet.delete_rows(2)

        # Duplicates
        workbook.create_sheet("Id doublons")
        worksheet = workbook["Id doublons"]
        if self.id['duplicates'] is not None:
            for row in dataframe_to_rows(self.id['duplicates']):
                worksheet.append(row)
        worksheet.delete_rows(2)

        # Enumerators
        workbook.create_sheet("Enquêteurs")
        worksheet = workbook["Enquêteurs"]
        if self.enumeratorissues['duration'] is not None:
            for row in dataframe_to_rows(self.enumeratorissues['duration']):
                worksheet.append(row)
        # worksheet.delete_rows(2)

        date_time = get_datetime()
        workbook.save(str(self.outputPath) + '/' + 'monitoring-' + date_time + '.xlsx')
        show_info_messagebox("Exportation réussie!!")

    def monitore(self):
        self.missingid()
        self.duplicatedid()
        self.end_before_start()
        self.lendayofsurvey()
        self.enumerators()
        self.export()


def get_datetime():
    """Retourne la date et l'heure pour nommer les fichiers"""
    current_datetime = datetime.datetime.now()
    # Format the new datetime
    formatted_datetime = current_datetime.strftime("%d%m%Y-%H%M%S")
    # Return the formatted datetime
    return formatted_datetime

