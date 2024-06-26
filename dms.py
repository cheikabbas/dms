import datetime
import os
import shutil

import openpyxl
import pandas as pd
import pyreadstat
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.stats import shapiro

from custom_widgets import ResultData
from messageBoxes import show_critical_messagebox, show_info_messagebox, show_warning_messagebox


class Dms:
    """Cette classe a des fonctions pour vérifier la qualité des données lors d'une collecte de données"""

    def __init__(self):
        self.duplicates = None
        self.na = None
        self.globalPath = None
        self.dataPath = None
        self.outputPath = None
        self.correctionsPath = None
        self.filesList = []
        self.data = None
        self.outliers = []
        self.sample_df = None
        self.workbook = openpyxl.Workbook()
        self.operations = ["#################### VERIFICATIONS A FAIRE ####################"]

    # Create new project
    def create_project(self, chemin, nom_projet, data):
        """Crée un projet de suivi de la qualité des données"""
        self.globalPath = os.path.join(chemin, nom_projet)
        if not os.path.exists(self.globalPath):
            os.makedirs(self.globalPath)
            file_config = os.path.join(self.globalPath, "config.dms")
            with open(file_config, "x") as f:
                f.write(f"Created on {get_datetime()}\n")
                f.write(self.globalPath)
            for i in ["data", "output", "corrections"]:
                os.makedirs(os.path.join(self.globalPath, i))
            self.dataPath = os.path.join(self.globalPath, "data")
            self.outputPath = os.path.join(self.globalPath, "output")
            self.correctionsPath = os.path.join(self.globalPath, "corrections")

            try:
                shutil.copy(data, self.dataPath)
            except IOError as e:
                show_critical_messagebox("Impossible de copier le fichier de données. " + e)

            return "Projet créé avec succès! Copier le fichier le données dans le dossier 'data' créé pour continuer."
        return "Dossier existe déjà."

    # Open existing project
    def open_project(self, chemin):
        self.globalPath = chemin
        self.dataPath = os.path.join(self.globalPath, "data")
        self.outputPath = os.path.join(self.globalPath, "output")
        self.correctionsPath = os.path.join(self.globalPath, "corrections")

    # Get Sheets names of Excel file
    @staticmethod
    def getsheetsnames(file):
        workbook = openpyxl.load_workbook(file)
        sheet_names = workbook.sheetnames
        return sheet_names

    # Load data file
    def load_data(self, fichier, sheet_name):
        self.data = pd.read_excel(fichier, sheet_name=sheet_name)

    def missing_count(self):
        nb_na = self.data.isnull().sum()
        pct_na = round(self.data.isna().mean() * 100, 2)
        self.na = pd.DataFrame({"NB Valeurs manquantes": nb_na, "% Valeurs manquantes": pct_na})
        self.na.insert(loc=0, column='Variable', value=self.na.index)
        five_pct = self.na[pct_na > 0.0].index.tolist()
        self.operations.append("### VARIABLES CONTENANT DES VALEURS MANQUANTES ###")
        for name in five_pct:
            op = f"-> [{name}]"
            self.operations.append(op)
        ResultData(self.na, "Valeurs manquantes")

    def duplicates_check(self, varnames):
        if len(varnames) > 0:
            self.duplicates = self.data[self.data.duplicated(subset=varnames, keep=False)]
            dups = self.duplicates.index.tolist()
            self.operations.append("### DOUBLONS ###")
            for name in dups:
                op = f"-> La ligne N°{name + 1} est un doublon si on considère la(les) variable(s) {varnames}."
                self.operations.append(op)
            ResultData(self.duplicates, "Doublons")

    def outliers_check(self, varname):
        if varname is not None:
            q1 = self.data[varname].quantile(0.25)
            q3 = self.data[varname].quantile(0.75)
            iqr = q3 - q1
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            self.outliers.append(self.data[(self.data[varname] < lower_bound) | (self.data[varname] > upper_bound)])
            self.outliers = pd.concat(self.outliers)
            self.operations.append("### VALEURS ABBERANTES ###")
            for name in self.outliers.index.tolist():
                op = f"-> La ligne N°{name + 1} a une valeur abberante pour la variable [{varname}]."
                self.operations.append(op)
            ResultData(self.outliers, "Valeurs aberrantes")

    def export_output(self):
        if (self.na is None) and (self.duplicates is None) and (len(self.outliers) == 0):
            show_info_messagebox("Aucun résultat à exporter")
        else:
            self.workbook.create_sheet("Valeurs manquantes")
            worksheet = self.workbook["Valeurs manquantes"]
            if self.na is not None:
                for row in dataframe_to_rows(self.na):
                    worksheet.append(row)
            worksheet.delete_rows(2)

            self.workbook.create_sheet("Doublons")
            worksheet = self.workbook["Doublons"]
            if self.duplicates is not None:
                for row in dataframe_to_rows(self.duplicates):
                    worksheet.append(row)
            worksheet.delete_rows(2)

            self.workbook.create_sheet("Valeurs abbérantes")
            worksheet = self.workbook["Valeurs abbérantes"]
            if len(self.outliers) > 0:
                for row in dataframe_to_rows(self.outliers, header=True):
                    worksheet.append(row)
            worksheet.delete_rows(2)

            self.write_corrections()
            try:
                date_time = get_datetime()
                self.workbook.save(str(self.outputPath) + '/' + 'output-' + date_time + '.xlsx')
                show_info_messagebox("Exportation réussie!!")
            except IOError:
                show_info_messagebox("Fermer le fichier output avant de recommencer")

    def reinit(self):
        self.na = None
        self.duplicates = None
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            self.workbook.remove(sheet)

    def write_corrections(self):
        date_time = get_datetime()
        correction_file = str(self.correctionsPath) + '/' + "corrections-" + date_time + ".txt"
        if os.path.exists(correction_file):
            with open(correction_file, 'a') as file:
                for text in self.operations:
                    file.write(text)
                    file.write('\n')
        else:
            with open(correction_file, 'w') as file:
                for text in self.operations:
                    file.write(text)
                    file.write('\n')

    def export_data(self, frmt):
        if frmt == "CSV":
            self.data.to_csv(self.dataPath + '/donnees.csv', index=False)
        if frmt == "Stata":
            self.data.to_stata(self.dataPath + '/donnees.dta', version=119)
        if frmt == "SPSS":
            pyreadstat.write_sav(self.data, self.dataPath + '/donnees.sav')

    def na_correction(self, method, var, fillvalue=None):
        if method == 'Supprimer':
            if var is not None:
                self.drop_na(var)
                show_info_messagebox(f"Données manquantes supprimées pour la variable {var}.")
        if method == "Remplacer":
            self.replace_na(var, fillvalue)
            show_info_messagebox(f"Données manquantes remplacées par {fillvalue} pour la variable {var}.")

    def drop_na(self, var):
        if var is not None:
            self.data.dropna(subset=[var], inplace=True)
        else:
            show_critical_messagebox("Indiquer la variable")

    def replace_na(self, var, fillvalue):
        if var is not None and fillvalue is not None:
            self.data[var].fillna(fillvalue, inplace=True)
        else:
            show_warning_messagebox("Indiquer la variable et la valeur de remplacement correctement.")

    def sample_data(self, n):
        if n is not None:
            self.sample_df = self.data.sample(n)
            show_info_messagebox(
                f"Echantillonnage éffectué avec succès. {n} lignes ont été tirées aléatoirement de la base de données.")

    def analyze(self, variables):
        statistics = {}

        for column in self.data[variables].columns:
            col_data = self.data[column]
            col_type = col_data.dtype

            if col_type == 'object':
                statistics[column] = {
                    'Variable': column,
                    'Nombre': col_data.count(),
                    'Unique': col_data.nunique(),
                    'Mode': col_data.mode().iloc[0],
                    'Fréquence': col_data.value_counts().max()
                }
            else:
                statistics[column] = {
                    'Variable': column,
                    'Nombre': col_data.count(),
                    'Moyenne': col_data.mean(),
                    'Ecart-type': col_data.std(),
                    'Min': col_data.min(),
                    '25%': col_data.quantile(0.25),
                    'Médiane': col_data.quantile(0.50),
                    '75%': col_data.quantile(0.75),
                    'Max': col_data.max()
                }

        return pd.DataFrame(statistics).T

    def test_statistique(self, var, test):
        res = ""
        if type(self.data[var]) != "object":
            try:
                if test == "Test de normalité":
                    statistic, p_value = shapiro(self.data[var])
                    res = f"<b>Statistique du test : </b>{statistic}<br><b>P-value : </b>{p_value}<br>"
                    if p_value < 0.05:
                        res = res + "La p_value est significative et donc la variable ne suit pas une loi normale."
                    else:
                        res = res + "La p-value n'est pas significative et donc la variable suit une loi normale."
            except ValueError:
                pass
        else:
            show_critical_messagebox("Le type de variable choisie ne permet d'éffectuer un test de normalité.")
        return res


def get_datetime():
    """Retourne la date et l'heure pour nommer les fichiers"""
    current_datetime = datetime.datetime.now()
    # Format the new datetime
    formatted_datetime = current_datetime.strftime("%d%m%Y-%H%M%S")
    # Return the formatted datetime
    return formatted_datetime
